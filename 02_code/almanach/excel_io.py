"""Excel (.xlsx) import / export of the full source hierarchy (CR-260704-1825-001).

Chosen over XML because a spreadsheet is far friendlier to hand-edit. The folder
hierarchy lives in a single **Folder** column as a `/`-separated path
(e.g. `AI / Generalist-Daily`); blank = ungrouped. **Row order = display order.**

Sources are identified by their **URL** (the real polling identity): a matched
URL updates its row in place (articles + read state survive a round-trip); a URL
present live but absent from the sheet is deleted (articles cascade); a new row
needs a URL. Upload REPLACES the whole hierarchy immediately — no staging, no
approval (the documented exception to the §11 two-zone safeguard; DATA_MODEL.md
§12).
"""
from __future__ import annotations

import io
import zipfile
from datetime import date
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

from . import db, models, settings_store
from .urls import canonical_source_url, is_valid_http_url

EXPORT_FILENAME_PREFIX = "almanach-sources-"
TEMPLATE_FILENAME = "almanach-sources-template.xlsx"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
SHEET_NAME = "Sources"
FOLDER_SEP = " / "

HEADERS = [
    "Folder", "Source name", "URL", "Feed URL", "Color",
    "Muted", "Reliability", "Impact", "Article count",
]
_COL_WIDTHS = [26, 26, 34, 34, 10, 8, 12, 10, 13]

# Header text (lowercased) → canonical key. Extra aliases make the parse
# forgiving of a user who renames or reorders columns.
_HEADER_KEYS = {
    "folder": "folder", "group": "folder", "folder path": "folder",
    "source name": "name", "name": "name", "source": "name", "provider": "name",
    "url": "url", "homepage": "url", "homepage url": "url",
    "feed url": "feed", "feed": "feed",
    "color": "color", "colour": "color",
    "muted": "muted",
    "reliability": "reliability",
    "impact": "impact",
    "article count": "count", "count": "count",
}

_MUTED_TRUE = {"yes", "true", "1", "x", "muted", "y"}


def export_filename() -> str:
    return f"{EXPORT_FILENAME_PREFIX}{date.today().isoformat()}.xlsx"


# ---------------- serialize: live DB -> .xlsx bytes ----------------


def _article_counts() -> dict:
    cur = db.get_connection().cursor()
    cur.execute("SELECT source_id, COUNT(*) AS n FROM article GROUP BY source_id")
    return {row["source_id"]: row["n"] for row in cur.fetchall()}


def _source_row(s: dict, folder_str: str, counts: dict) -> list:
    return [
        folder_str,
        s["display_name"],
        s["url"],
        s["feed_url"],
        s["colour"],
        "yes" if s["muted"] else "no",
        s["reliability"],
        s["impact"],
        counts.get(s["id"], 0),
    ]


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def serialize_to_xlsx() -> bytes:
    """Serialise the live folder tree + sources to a one-sheet workbook, rows in
    sidebar (display) order. Empty folders emit a folder-only row so they survive
    the round-trip."""
    folders = models.list_folders()
    sources = models.list_sources()
    counts = _article_counts()

    children_of: dict = {}
    for f in folders:
        children_of.setdefault(f["parent_id"], []).append(f)
    sources_of: dict = {}
    for s in sources:
        sources_of.setdefault(s.get("folder_id"), []).append(s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    def emit(node: dict, parent_path: list) -> None:
        path = parent_path + [node["name"]]
        folder_str = FOLDER_SEP.join(path)
        srcs = sources_of.get(node["id"], [])
        kids = children_of.get(node["id"], [])
        if not srcs and not kids:
            ws.append([folder_str] + [""] * (len(HEADERS) - 1))
        for s in srcs:
            ws.append(_source_row(s, folder_str, counts))
        for child in kids:
            emit(child, path)

    for root in children_of.get(None, []):
        emit(root, [])
    for s in sources_of.get(None, []):
        ws.append(_source_row(s, "", counts))

    _style_header(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- template (virgin workbook) ----------------


def build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for row in [
        ["AI / Generalist-Daily", "Provider one", "https://example.com",
         "https://example.com/feed", "#378ADD", "no", "high", "medium", 0],
        ["AI / Generalist-Daily", "Provider two", "https://example.org",
         "", "", "no", "medium", "medium", 0],
        ["AI", "Provider three (directly under the group)", "https://example.net",
         "", "", "no", "medium", "medium", 0],
        ["", "Unsorted provider", "https://example.info",
         "", "", "no", "low", "low", 0],
    ]:
        ws.append(row)
    _style_header(ws)

    guide = wb.create_sheet("How to use")
    lines = [
        "ALMANACH — sources template",
        "",
        "1. Each row is one news provider. Fill in the columns on the 'Sources' tab.",
        "2. Folder  : the group path, e.g. 'AI / Generalist-Daily'. Blank = ungrouped.",
        "             Use ' / ' to nest a subgroup inside a group (up to 5 levels).",
        "3. Row order = the order shown in the sidebar. Rearrange rows to reorder.",
        "4. Re-upload this file — it REPLACES your whole list immediately, with no",
        "   confirmation. Providers missing from the file are removed (with their",
        "   articles). A source is matched by its URL, so keep the URL to keep its",
        "   articles; change the URL and it becomes a new provider.",
        "",
        "COLUMNS",
        "  Source name  display name in the sidebar (defaults to the URL host if blank)",
        "  URL          homepage URL — REQUIRED for every provider",
        "  Feed URL     RSS/Atom/sitemap URL; defaults to the homepage URL if blank",
        "  Color        hex dot colour, e.g. #378ADD; a palette colour is picked if blank",
        "  Muted        yes / no (default no)",
        "  Reliability  high / medium / low (default medium)",
        "  Impact       high / medium / low (default medium)",
        "  Article count  informational only — ignored on import",
    ]
    for i, line in enumerate(lines, start=1):
        guide.cell(row=i, column=1, value=line)
    guide.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- parse: .xlsx bytes -> ordered rows ----------------


def _parse_rating(raw: str) -> str:
    return raw if raw in models.RATING_VALUES else "medium"


def parse_xlsx(data: bytes) -> dict:
    """Parse workbook bytes to {"rows": [ {folder_path, source|None}, ... ]} in
    sheet order. Raises ValueError with a user-facing message on any problem."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, OSError, KeyError) as e:
        raise ValueError(
            "This file is not a valid Excel (.xlsx) workbook. Use the blank template."
        ) from e

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError("The workbook is empty — no header row found.")

    idx: dict[str, int] = {}
    for i, cell in enumerate(header or []):
        if cell is None:
            continue
        key = _HEADER_KEYS.get(str(cell).strip().lower())
        if key and key not in idx:
            idx[key] = i
    if "url" not in idx and "name" not in idx and "folder" not in idx:
        raise ValueError(
            "Missing the expected columns (Folder, Source name, URL…). "
            "Download the blank template and use its headings."
        )

    def cell(raw, key: str) -> str:
        i = idx.get(key)
        if i is None or i >= len(raw) or raw[i] is None:
            return ""
        return str(raw[i]).strip()

    out_rows: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue
        folder = cell(raw, "folder")
        name = cell(raw, "name")
        url = cell(raw, "url")
        if not folder and not name and not url:
            continue  # blank row
        path = [p.strip() for p in folder.split("/") if p.strip()] if folder else []
        if len(path) > db.MAX_FOLDER_DEPTH:
            raise ValueError(
                f"Folder path '{folder}' nests deeper than {db.MAX_FOLDER_DEPTH} "
                "levels — flatten it and retry."
            )
        source = None
        if name or url:
            source = {
                "name": name,
                "url": url,
                "feed": cell(raw, "feed"),
                "color": cell(raw, "color"),
                "muted": cell(raw, "muted").lower() in _MUTED_TRUE,
                "reliability": _parse_rating(cell(raw, "reliability").lower()),
                "impact": _parse_rating(cell(raw, "impact").lower()),
            }
        out_rows.append({"folder_path": path, "source": source})
    return {"rows": out_rows}


# ---------------- replace: ordered rows -> live DB ----------------


def replace_hierarchy(parsed: dict) -> dict:
    """Replace the entire live folder tree + source list from parsed rows.

    One IMMEDIATE transaction — a failed import leaves the DB untouched. Sources
    match by canonical URL and update in place (articles survive); unmatched live
    sources are deleted (articles cascade); rows without a valid URL are skipped
    and reported."""
    rows = parsed["rows"]

    # Palette colours resolve BEFORE the replace transaction — settings_store
    # runs its own transaction and db.transaction is not re-entrant.
    for r in rows:
        s = r["source"]
        if s and not s["color"]:
            s["color"] = settings_store.next_palette_colour()

    skipped: list = []
    removed = 0
    applied = 0
    groups = len({r["folder_path"][0] for r in rows if r["folder_path"]})

    with db.transaction() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, url FROM source")
        by_url = {row["url"]: row["id"] for row in cur.fetchall()}

        # --- folders: full rebuild (source.folder_id -> SET NULL via FK) ---
        cur.execute("DELETE FROM folder")
        folder_id_by_path: dict[tuple, str] = {}
        sibling_pos: dict = {}

        def ensure_folder(path_tuple: tuple):
            if not path_tuple:
                return None
            if path_tuple in folder_id_by_path:
                return folder_id_by_path[path_tuple]
            parent_id = ensure_folder(path_tuple[:-1])
            pos = sibling_pos.get(parent_id, 0)
            sibling_pos[parent_id] = pos + 1
            fid = db.new_id()
            cur.execute(
                "INSERT INTO folder (id, parent_id, name, position, collapsed, "
                "muted, depth, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (fid, parent_id, path_tuple[-1], float(pos), 0, 0,
                 len(path_tuple), db.now_iso()),
            )
            folder_id_by_path[path_tuple] = fid
            return fid

        # Materialise every folder first, in row order, so display order holds.
        for r in rows:
            ensure_folder(tuple(r["folder_path"]))

        # --- sources: match/update, insert, then delete the rest ---
        kept: set = set()
        claimed: set = set()
        for position, r in enumerate(rows):
            s = r["source"]
            if s is None:
                continue
            canonical = (
                canonical_source_url(s["url"])
                if s["url"] and is_valid_http_url(s["url"])
                else ""
            )
            if not canonical:
                skipped.append(s["name"] or "(unnamed)")
                continue
            if canonical in claimed:
                skipped.append(s["name"] or canonical)  # duplicate url in file
                continue
            folder_id = (
                folder_id_by_path.get(tuple(r["folder_path"]))
                if r["folder_path"] else None
            )
            name = s["name"] or urlparse(canonical).hostname or canonical
            match_id = by_url.get(canonical)
            if match_id and match_id not in kept:
                cur.execute(
                    "UPDATE source SET display_name = ?, "
                    "feed_url = COALESCE(NULLIF(?, ''), feed_url), colour = ?, "
                    "muted = ?, reliability = ?, impact = ?, folder_id = ?, "
                    "position = ? WHERE id = ?",
                    (name, s["feed"], s["color"], int(s["muted"]),
                     s["reliability"], s["impact"], folder_id,
                     float(position), match_id),
                )
                kept.add(match_id)
            else:
                sid = db.new_id()
                cur.execute(
                    "INSERT INTO source (id, url, feed_url, discovery_method, "
                    "display_name, colour, muted, created_at, folder_id, "
                    "position, reliability, impact) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (sid, canonical, s["feed"] or canonical,
                     "alternate_link" if s["feed"] else "common_path",
                     name, s["color"], int(s["muted"]), db.now_iso(),
                     folder_id, float(position), s["reliability"], s["impact"]),
                )
                kept.add(sid)
            claimed.add(canonical)
            applied += 1

        for url, sid in by_url.items():
            if sid not in kept:
                cur.execute("DELETE FROM source WHERE id = ?", (sid,))
                removed += 1

    return {
        "sources": applied,
        "groups": groups,
        "removed": removed,
        "skipped": skipped,
    }
